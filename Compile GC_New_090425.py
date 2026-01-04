# -*- coding: utf-8 -*-
"""
Updated Sept 4, 2025
Author: tpham + ChatGPT
"""

import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#%% ===================== CONFIGURATION ===================== #
MASTER_FOLDER = r"Y:\5900\HydrogenTechFuelCellsGroup\CO2R\Nhan P\Experiments\CO2 Cell Testing\TS2\2NP50_Conditioning 2 slpm\GC Data"

ANALYSIS_GASES = [
    {"name": "Carbon Monoxide", "Chan#": 1, "range": (60, 90)},
    {"name": "Hydrogen",        "Chan#": 1, "range": (21, 26)},
    {"name": "Carbon Monoxide", "Chan#": 2, "range": (60, 90)},
    {"name": "Carbon Dioxide", "Chan#": 3, "range": (20, 26)}
]

IDENTIFIER_GASES = {
    "N2": {"name": "Nitrogen", "Chan#": 3, "range": (15, 20)},
    "CO2": {"name": "Carbon Dioxide", "Chan#": 3, "range": (20, 26)}
}

GROUP_COLORS = {
    "Cathode": "C6EFCE",
    "QC":      "FFEB9C",
    "N2":      "D9D9D9",
    "N/A":     "FFC7CE"
}

PASS_FILL = "C6EFCE"   # Green
FAIL_FILL = "FFC7CE"   # Red

#%% ======================================================== #
def get_txt_table_data(file_path):
    try:
        with open(file_path, 'r') as f:
            lines = f.readlines()
        start_index = None
        end_index = None
        for i, line in enumerate(lines):
            if '>==CT==' in line:
                start_index = i
            if 'Totals' in line:
                end_index = i
                break
        if start_index is not None and end_index is not None:
            table_lines = lines[start_index+1:end_index]
        else:
            print("Table section not found:", file_path)
            return pd.DataFrame()

        data = []
        headers = []
        for line in table_lines:
            stripped = line.strip()
            if stripped == '':
                continue
            parts = [p.strip() for p in line.split('\t')]
            if not headers:
                headers = parts
            else:
                data.append(parts)
        if not data:
            print(f"No valid data found in file: {file_path}")
            return pd.DataFrame()
        return pd.DataFrame(data, columns=headers)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return pd.DataFrame()

def find_gas(df, gas):
    try:
        df['Retention'] = pd.to_numeric(df['Retention'], errors='coerce')
        df['Chan#'] = pd.to_numeric(df['Chan#'], errors='coerce', downcast='integer')
    except Exception as e:
        print(f"[{gas['name']}] Conversion error: {e}")
        return None
    matches = df[(df['Chan#']==gas['Chan#']) & (df['Retention'].between(gas['range'][0], gas['range'][1]))]
    if matches.empty:
        print(f"[{gas['name']}] No match")
        return None
    if len(matches) > 1:
        print(f"[{gas['name']}] WARNING: multiple matches")
        return matches.iloc[0].to_dict()
    return matches.iloc[0].to_dict()

def get_group_id(df, folder_name):
    g_co2 = find_gas(df, IDENTIFIER_GASES["CO2"])
    g_n2  = find_gas(df, IDENTIFIER_GASES["N2"])
    print(f"\n[{folder_name}] Identifier Gases:")
    print(f"  CO2: {g_co2}")
    print(f"  N2:  {g_n2}")
    try:
        if g_co2 is not None:
            estd = float(g_co2['ESTD'])
            if estd < 0.4:
                return "N2"
            elif 0.4 <= estd <= 1.0:
                return "QC"
            else:
                return "Cathode"
        else:
            if g_n2 is not None:
                return "N2"
            else:
                return "N/A"
    except Exception as e:
        print(f"[{folder_name}] Error in ESTD logic: {e}")
        return "Error"

def extract_analysis_results(df, group_id):
    results = {}
    groups = ["Cathode","QC","N2","N/A"]
    for g in groups:
        for gas in ANALYSIS_GASES:
            label = f"{g}_{gas['name'].replace(' ','_')}_Chan{gas['Chan#']}"
            if g == group_id:
                match = find_gas(df, gas)
                results[label] = None if match in [None,'error'] else match['ESTD']
            else:
                results[label] = "N/A"
    return results

def extract_timestamp(name):
    m = re.search(r'(\d{8}T\d{6})', name)
    return m.group(1) if m else None

#%% ======================================================== #
def main():
    data_rows = []
    for folder in os.listdir(MASTER_FOLDER):
        folder_path = os.path.join(MASTER_FOLDER, folder)
        if not os.path.isdir(folder_path):
            continue
        file_path = os.path.join(folder_path, "SAMPRSLT.TXT")
        if not os.path.exists(file_path):
            continue
        df = get_txt_table_data(file_path)
        if df.empty:
            continue
        group_id = get_group_id(df, folder)
        row_data = extract_analysis_results(df, group_id) if group_id != "Error" else {}
        timestamp = extract_timestamp(folder)
        row = {"FolderName": folder, "Group ID": group_id, "Timestamp": timestamp, **row_data}
        data_rows.append(row)

    final_df = pd.DataFrame(data_rows)
    final_df = final_df.sort_values(by='Timestamp')

    # Reorder columns: GROUP → GAS
    ordered_cols = ["FolderName","Group ID","Timestamp"]
    groups_order = ["Cathode","QC","N2","N/A"]
    for g in groups_order:
        for gas in ANALYSIS_GASES:
            col = f"{g}_{gas['name'].replace(' ','_')}_Chan{gas['Chan#']}"
            if col in final_df.columns:
                ordered_cols.append(col)
    final_df = final_df[ordered_cols]

    # Cathode-only sheet with Pass/Fail based on most recent QC
    cathode_rows = []
    last_qc_value = None
    for _, row in final_df.iterrows():
        if row["Group ID"] == "QC":
            for col in row.index:
                if col.startswith("QC_Carbon_Dioxide"):
                    try:
                        last_qc_value = float(row[col])
                    except:
                        last_qc_value = None
        elif row["Group ID"] == "Cathode":
            cath_row = row.copy()
            if last_qc_value is not None and 0.4 <= last_qc_value <= 1.0:
                cath_row["Pass/Fail"] = "PASS"
            else:
                cath_row["Pass/Fail"] = f"FAIL (QC={last_qc_value})"
            cathode_rows.append(cath_row)
    cathode_df = pd.DataFrame(cathode_rows)

    # Export to Excel
    out_file = os.path.join(MASTER_FOLDER, "Grouped_Analysis.xlsx")
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="All Data", index=False)
        cathode_df.to_excel(writer, sheet_name="Cathode Only", index=False)

    # Color coding
    wb = load_workbook(out_file)
    ws_all = wb["All Data"]
    ws_cath = wb["Cathode Only"]

    # Color All Data columns
    for idx, col in enumerate(final_df.columns, start=1):
        for group, color in GROUP_COLORS.items():
            if col.startswith(group):
                for cell in ws_all[ws_all.cell(row=1, column=idx).column_letter]:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                break

    # Color Pass/Fail in Cathode sheet
    pass_col_idx = None
    for idx, cell in enumerate(ws_cath[1], start=1):
        if cell.value == "Pass/Fail":
            pass_col_idx = idx
            break
    if pass_col_idx:
        for row in ws_cath.iter_rows(min_row=2, min_col=pass_col_idx, max_col=pass_col_idx):
            for cell in row:
                if str(cell.value).startswith("PASS"):
                    cell.fill = PatternFill(start_color=PASS_FILL, end_color=PASS_FILL, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=FAIL_FILL, end_color=FAIL_FILL, fill_type="solid")

    wb.save(out_file)
    print("Excel file created successfully with corrected column order and color coding:", out_file)


if __name__ == "__main__":
    main()
